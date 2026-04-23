#!/usr/bin/env python3
"""Build an xlsx of all signed-up practices, grouped by post-April-2026 ICB.

Practice universe matches the dashboard map (src/components/StatsPanel.jsx):
iterate practices_geocoded.json and keep any practice whose ODS is in
waitlist, live_customers or live_customers_full_planner.

Status tiers (src/components/DashboardMap.jsx):
  Live       -> ODS in live_customers_full_planner.json
  Onboarding -> ODS in live_customers.json but not in _full_planner
  Signed up  -> ODS in waitlist_ods.json (HubSpot list 1535)

ICB column uses the post-2026-04-01 name. Non-merging ICBs pass through.
Merging ICBs are relabelled via scripts/icb_mapper.py:
 - 9 simple 1:1 relabels
 - 2 SICBL-split ICBs (Herts & W Essex, Suffolk & NE Essex) → NHS ODS API
 - Frimley split → per-practice xlsx move table
"""
import json
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from icb_mapper import (
    SicblCache,
    UnresolvableSplit,
    build_frimley_map,
    resolve_icb,
)

ROOT = Path(__file__).parent.parent
DATA = ROOT / "public" / "data"
ODS_XLSX = ROOT / "ODS+Change+Summary+ICB+Mergers+Phase+1+Apr+2026 (2).xlsx"
SICBL_CACHE = Path(__file__).parent / ".sicbl_cache.json"

waitlist = set(json.load(open(DATA / "waitlist_ods.json")))
live_all = set(json.load(open(DATA / "live_customers.json")))
live_full = set(json.load(open(DATA / "live_customers_full_planner.json")))
practices = json.load(open(DATA / "practices_geocoded.json"))


def status_for(ods: str):
    if ods in live_full: return "Live"
    if ods in live_all:  return "Onboarding"
    if ods in waitlist:  return "Signed up"
    return None


def main():
    frimley_map = build_frimley_map(ODS_XLSX)
    sicbl = SicblCache(SICBL_CACHE)

    rows = []
    errors = []
    for p in practices:
        ods = p["ods"].upper()
        st = status_for(ods)
        if not st:
            continue
        pre_icb = (p.get("icb") or "").strip()
        try:
            post_icb = resolve_icb(pre_icb, ods, sicbl_lookup=sicbl, frimley_map=frimley_map)
        except UnresolvableSplit as e:
            errors.append(str(e))
            post_icb = f"UNRESOLVED: {pre_icb}"
        changed = (post_icb != pre_icb) and not post_icb.startswith("UNRESOLVED")
        rows.append({
            "icb":      post_icb,
            "pre_icb":  pre_icb or "(unknown)",
            "changed":  "Yes" if changed else "No",
            "pcn":      p.get("pcn_name") or "",
            "pcn_code": p.get("pcn_code") or "",
            "ods":      ods,
            "name":     p.get("name") or "",
            "status":   st,
            "patients": p.get("patients") or 0,
        })

    rows.sort(key=lambda r: (r["icb"], r["pcn"] or "zzz", r["name"]))

    print(f"Total signed practices: {len(rows)}")
    by_status = defaultdict(int)
    for r in rows:
        by_status[r["status"]] += 1
    print(f"  Live: {by_status['Live']}  Onboarding: {by_status['Onboarding']}  Signed up: {by_status['Signed up']}")
    print(f"  ICB renamed by merger: {sum(1 for r in rows if r['changed']=='Yes')}")
    if errors:
        print(f"\n[!] {len(errors)} unresolvable split practices:")
        for e in errors:
            print("   ", e)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sign-ups by ICB (post-Apr 2026)"

    headers = [
        "ICB (post-2026-04-01)", "Current ICB", "ICB changing?",
        "Primary Care Network (PCN)", "PCN code",
        "ODS code", "Practice name", "Status", "Patients",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E2A4A")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    status_colors = {
        "Live":       "15803D",
        "Onboarding": "22C55E",
        "Signed up":  "F59E0B",
    }

    for r in rows:
        ws.append([
            r["icb"], r["pre_icb"], r["changed"],
            r["pcn"], r["pcn_code"],
            r["ods"], r["name"], r["status"], r["patients"],
        ])
        row_idx = ws.max_row
        sc = status_colors.get(r["status"])
        if sc:
            cell = ws.cell(row=row_idx, column=8)
            cell.fill = PatternFill("solid", fgColor=sc)
            cell.font = Font(color="FFFFFF", bold=True)

    widths = {1: 48, 2: 40, 3: 14, 4: 42, 5: 10, 6: 10, 7: 48, 8: 12, 9: 10}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    summary = wb.create_sheet("Summary by ICB")
    summary.append(["ICB (post-2026-04-01)", "Live", "Onboarding", "Signed up", "Total"])
    for cell in summary[1]:
        cell.font = header_font
        cell.fill = header_fill

    agg = defaultdict(lambda: {"Live": 0, "Onboarding": 0, "Signed up": 0})
    for r in rows:
        agg[r["icb"]][r["status"]] += 1

    for icb in sorted(agg):
        a = agg[icb]
        total = sum(a.values())
        summary.append([icb, a["Live"], a["Onboarding"], a["Signed up"], total])

    summary.append([
        "TOTAL",
        sum(a["Live"] for a in agg.values()),
        sum(a["Onboarding"] for a in agg.values()),
        sum(a["Signed up"] for a in agg.values()),
        sum(sum(a.values()) for a in agg.values()),
    ])
    for cell in summary[summary.max_row]:
        cell.font = Font(bold=True)

    summary.column_dimensions["A"].width = 60
    for c in "BCDE":
        summary.column_dimensions[c].width = 12
    summary.freeze_panes = "A2"

    out = ROOT / "signups_by_icb.xlsx"
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"  Detail rows: {len(rows)}  |  ICBs: {len(agg)}")

    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
